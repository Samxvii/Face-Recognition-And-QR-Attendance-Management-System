from pathlib import Path
from tkinter import Canvas, Button, PhotoImage,ttk
import tkinter as tk
from tkinter import *
from PIL import ImageTk,Image
import os
import cv2
import openpyxl
import random
import qrcode

OUTPUT_PATH = Path(__file__).parent
ASSETS_PATH = OUTPUT_PATH / Path(r"C:\Users\sansk\OneDrive\Desktop\New folder\final\build\assets\frame0")
def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)

window = tk.Tk()
window.title("Face Recognition And QR Based Attendance Management System") 
window.geometry("1115x677")
style=ttk.Style(window)

def sam():
    sam="C:\\Users\\sansk\\OneDrive\\Desktop\\New folder\\app\\dist\\AMS_Run.exe"
    os.system('"%s"'% sam)

wb = openpyxl.load_workbook("C:\\Users\\sansk\\OneDrive\\Desktop\\New folder\\final\\build\\StudentDetails.xlsx")
sheet = wb.active

def generate_qr_code():
    row = random.randint(1, sheet.max_row)
    data = sheet.cell(row=row, column=1).value
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img_with_border = Image.new("RGB", (img.size[0] + 2, img.size[1] + 2), color="white")
    img_with_border.paste(img, (1, 1))
    photo = ImageTk.PhotoImage(img_with_border)
    qr_code_label.configure(image=photo)
    qr_code_label.image = photo
    window.after(5000, generate_qr_code)

def open_csv():
   open_csv="C:\\Users\\sansk\\OneDrive\\Desktop\\New folder\\faltu\\faltu.exe"
   os.system('"%s"'% open_csv)

canvas = Canvas(
    window,
    height = 677,
    width = 1115,
    bd = 0,
    highlightthickness = 0,
    relief = "ridge"
)

canvas.place(x = 0, y = 0)

cap = cv2.VideoCapture(0)

bg_image = Image.open("C:\\Users\\sansk\\OneDrive\\Desktop\\New folder\\final\\build\\assets\\frame0\\5.png") 
bg_photo = ImageTk.PhotoImage(bg_image)
bg_item = canvas.create_image(0, 0, image=bg_photo, anchor='nw')

cap = cv2.VideoCapture(0)
video_item = None

face_cascade = cv2.CascadeClassifier('C:/Users/sansk/AppData/Local/Programs/Python/Python311/Lib/site-packages/cv2/data/haarcascade_frontalface_default.xml')

def adjust_camera_window_size():
    global video_item 
    canvas.delete(video_item) 
    ret, frame = cap.read()
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))
    for (x, y, w, h) in faces:
        cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
    new_width = 450
    new_height = 300
    resized_frame = cv2.resize(frame, (new_width, new_height))
    resized_frame = cv2.cvtColor(resized_frame, cv2.COLOR_BGR2RGB)
    resized_image = Image.fromarray(resized_frame)
    resized_photo = ImageTk.PhotoImage(resized_image)
    video_item = canvas.create_image(600, 200, image=resized_photo, anchor='nw')
    canvas.itemconfig(video_item, image=resized_photo)
    canvas.image = resized_photo 
    window.after(10, adjust_camera_window_size)
adjust_camera_window_size()

canvas.pack(fill= "both", expand=True)
canvas.create_rectangle(
    10.0,
    11.0,
    1101.0,
    158.0,
    fill="#749BFF",
    outline="")

canvas.create_text(
    29.0,
    21.0,
    anchor="nw",
    text="             Face Recognition And QR Based \n             Attendance Management System",
    fill="#0038FF",
    font=("Inter Bold", 50 * -1)
)

canvas.create_rectangle(
    552.0,
    180.0,
    558.0,
    653.0000305175781,
    fill="#000000",
    outline="")
    
button_image_1 = PhotoImage(
    file=relative_to_assets("button_1.png"))
button_1 = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    command=sam,
    relief="flat"
)
button_1.place(
    x=667.0,
    y=531.0,
    width=338,
    height=47
)
button_image_2 = PhotoImage(
    file=relative_to_assets("button_2.png"))
button_2 = Button(
    image=button_image_2,
    borderwidth=0,
    highlightthickness=0,
    command=open_csv,
    relief="flat"
)
button_2.place(
    x=283.0,
    y=531.0,
    width=135.0,
    height=80.0
)
qr_code_label=tk.Label(window,text='QR to display here')
canvas.create_rectangle(
    124.0,
    201.0,
    424.0,
    501.0,
    fill="#FFFFFF",
    outline="")
canvas.create_window( 275,355,  window=qr_code_label)

button_image_3 = PhotoImage(
    file=relative_to_assets("button_3.png"))
button_3 = Button(
    image=button_image_3,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: generate_qr_code(),
    relief="flat"
)
button_3.place(
    x=124.0,
    y=531.0,
    width=142,
    height=85
)
window.resizable(False, False)
window.mainloop()
